"""Importador de Excel: archivo .xlsx subido a mano -> SQLite.

Reemplaza el enfoque anterior (Microsoft Graph + Azure AD app registration)
por algo mucho mas simple: cualquier persona con el archivo actualizado lo
sube desde el dashboard, se parsea aqui y todos los demas usuarios ven los
cambios en su siguiente refresco automatico (ver hx-trigger en el template).
"""
import hashlib
import io
import logging
from datetime import date, datetime

import openpyxl

from app import db
from app.config import SHEETS

logger = logging.getLogger("dashboard.excel_import")


class ExcelImportError(ValueError):
    """Se lanza cuando el archivo subido no tiene la forma esperada."""


# Los encabezados normalizados del Excel no siempre coinciden con los nombres
# de columna que usamos en SQLite (ver app/db.py TABLE_COLUMNS). Este mapa
# traduce solo donde hace falta; cualquier encabezado no listado se usa tal cual.
HEADER_ALIASES = {
    "roadmap": {
        "roadmap_task": "task",
        "roadmap_start": "start_date",
        "roadmap_end": "end_date",
        "roadmap_%": "pct",
    },
}


def _normalize_header(cell) -> str:
    return str(cell or "").strip().lower().replace(" ", "_")


def _clean_cell(value):
    """Excel guarda fechas como datetime (con hora, aunque sea 00:00:00) --
    a nadie en el dashboard le importa la hora, asi que aqui se recorta a
    solo la fecha (YYYY-MM-DD) para las 3 hojas por igual.
    """
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


def _known_headers_for(table: str) -> set[str]:
    """Nombres de encabezado (ya normalizados, antes de aplicar alias) que
    reconocemos para esta tabla -- se usan para detectar en que fila esta
    el encabezado real dentro del Excel (puede no ser la fila 0 si hay
    titulos/filas en blanco antes, como pasa con el workbook real).
    """
    aliases = HEADER_ALIASES.get(table, {})
    return set(aliases.keys()) | set(db.TABLE_COLUMNS[table])


def _find_header_row(values: list[list], table: str) -> int | None:
    """Busca la fila con mas coincidencias contra los encabezados conocidos.
    Requiere al menos 2 coincidencias para evitar falsos positivos con
    filas de titulo que por casualidad tengan una palabra parecida.
    """
    known = _known_headers_for(table)
    best_row, best_score = None, 0
    for i, row in enumerate(values):
        score = sum(1 for cell in row if _normalize_header(cell) in known)
        if score > best_score:
            best_row, best_score = i, score
    return best_row if best_score >= 2 else None


def _rows_from_values(values: list[list], table: str) -> list[dict]:
    """Detecta automaticamente la fila de encabezados (puede no ser la 0,
    ver docstring de _find_header_row) y las columnas con datos reales
    (ignora columnas en blanco a la izquierda/derecha del rango usado).
    """
    if not values:
        return []

    header_idx = _find_header_row(values, table)
    if header_idx is None:
        raise ExcelImportError(
            f"No pude encontrar la fila de encabezados en la hoja de '{table}'. "
            f"Revisa que tenga columnas como: {', '.join(sorted(db.TABLE_COLUMNS[table]))}"
        )

    aliases = HEADER_ALIASES.get(table, {})
    header_row = values[header_idx]
    # Solo nos quedamos con las columnas que tienen un encabezado no vacio --
    # asi ignoramos columnas en blanco fuera del rango real de datos.
    col_map: dict[int, str] = {}
    for col_idx, cell in enumerate(header_row):
        normalized = _normalize_header(cell)
        if normalized:
            col_map[col_idx] = aliases.get(normalized, normalized)

    rows = []
    for raw_row in values[header_idx + 1:]:
        row = {name: (_clean_cell(raw_row[col_idx]) if col_idx < len(raw_row) else "") for col_idx, name in col_map.items()}
        # Saltar filas totalmente vacias (comunes al final de un rango en Excel).
        if any(str(v).strip() for v in row.values() if v is not None):
            rows.append(row)
    return rows


def import_workbook(file_bytes: bytes, filename: str) -> dict:
    """Parsea el .xlsx subido y reemplaza el contenido de las 3 tablas en SQLite.

    Tambien detecta si el archivo subido es byte-por-byte identico al ultimo
    que se importo -- comun cuando alguien vuelve a subir una copia local
    vieja creyendo que ya tiene los cambios mas recientes de SharePoint.
    """
    file_hash = hashlib.sha256(file_bytes).hexdigest()
    previous_hash = db.get_sync_meta("last_upload_hash")
    is_duplicate = file_hash == previous_hash

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception as exc:
        raise ExcelImportError(f"No se pudo abrir el archivo como Excel valido: {exc}") from exc

    missing_sheets = [name for name in SHEETS.values() if name not in wb.sheetnames]
    if missing_sheets:
        raise ExcelImportError(
            f"Al workbook le faltan las hojas: {', '.join(missing_sheets)}. "
            f"Hojas encontradas: {', '.join(wb.sheetnames)}"
        )

    # Metadata interna del propio archivo (la pone Excel/SharePoint al guardar).
    # Util para que el usuario compare visualmente si subio la version correcta.
    file_modified = wb.properties.modified or wb.properties.created

    imported_counts = {}
    for table, sheet_name in SHEETS.items():
        ws = wb[sheet_name]
        values = [list(row) for row in ws.iter_rows(values_only=True)]
        rows = _rows_from_values(values, table)
        db.replace_table(table, rows)
        imported_counts[table] = len(rows)

    db.set_sync_meta("last_upload_filename", filename)
    db.set_sync_meta("last_upload_hash", file_hash)
    db.set_sync_meta("source_file_modified", str(file_modified) if file_modified else "")
    db.mark_synced_now()
    logger.info("Import OK desde '%s': %s (duplicado=%s)", filename, imported_counts, is_duplicate)
    return {
        "status": "imported",
        "filename": filename,
        "counts": imported_counts,
        "is_duplicate": is_duplicate,
        "file_modified": str(file_modified) if file_modified else None,
    }
